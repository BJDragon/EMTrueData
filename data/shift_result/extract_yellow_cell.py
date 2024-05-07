import openpyxl
from openpyxl.styles import PatternFill

# 打开已有的xlsx文件
wb = openpyxl.load_workbook('shift_result_data.xlsx')
ws = wb.active  # 激活当前活跃的Worksheet
print(ws)
# 创建新的Workbook和Worksheet
new_wb = openpyxl.Workbook()
new_ws = new_wb.active

# 检查黄色背景的RGB值，Excel中黄色通常是'FFFF00'
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# 逐行逐列检查
for row in ws.iter_rows():
    for cell in row:
        if cell.fill.fgColor.index == 'FFFFFF00':
            print(cell.value)
            # 如果单元格背景是黄色，复制该值到新表格
            # 获取当前单元格的位置
            new_ws_coordinate = 'A'+str(cell.row)
            new_cell = new_ws[new_ws_coordinate]
            new_cell.value = cell.value
            # new_cell.fill = yellow_fill  # 也设置为黄色背景

# 保存新的excel文件
new_wb.save('shift_goal_value.xlsx')
